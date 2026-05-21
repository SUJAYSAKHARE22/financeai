"""
AI_CA Report Generator
Generates PDF and Excel reports from financial data.
Uses only stdlib + reportlab + openpyxl (no heavy deps).
"""
import io
import json
from datetime import datetime
from typing import Optional


def _try_reportlab():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        return True
    except ImportError:
        return False


def _try_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        return False


def generate_pdf_report(report_data: dict, report_type: str) -> bytes:
    """Generate PDF report. Returns bytes."""
    if not _try_reportlab():
        # Fallback: return a plain-text PDF substitute as bytes
        return _plain_text_report(report_data, report_type).encode()

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title = f"FinanceAI — {report_type.upper()} Report"
    story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}", styles["Normal"]))
    story.append(Spacer(1, 20))

    def _add_section(heading: str, data: dict):
        story.append(Paragraph(f"<b>{heading}</b>", styles["Heading2"]))
        rows = [["Item", "Value"]]
        for k, v in data.items():
            if isinstance(v, float):
                v = f"₹{v:,.2f}"
            elif isinstance(v, dict):
                continue
            rows.append([str(k).replace("_", " ").title(), str(v)])
        t = Table(rows, colWidths=[260, 200])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdf4")]),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 16))

    # Render sections
    for section_name, section_data in report_data.items():
        if isinstance(section_data, dict):
            _add_section(section_name.replace("_", " ").title(), section_data)
        elif isinstance(section_data, list) and section_data and isinstance(section_data[0], str):
            story.append(Paragraph(f"<b>{section_name.replace('_', ' ').title()}</b>", styles["Heading2"]))
            for item in section_data:
                story.append(Paragraph(f"• {item}", styles["Normal"]))
            story.append(Spacer(1, 12))

    doc.build(story)
    return buffer.getvalue()


def generate_excel_report(report_data: dict, report_type: str) -> bytes:
    """Generate Excel report. Returns bytes."""
    if not _try_openpyxl():
        return _plain_text_report(report_data, report_type).encode()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.upper()

    header_fill = PatternFill("solid", fgColor="10b981")
    alt_fill = PatternFill("solid", fgColor="f0fdf4")
    bold = Font(bold=True, color="FFFFFF")
    header_font = Font(bold=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row = 1
    ws.cell(row, 1, f"FinanceAI — {report_type.upper()} Report").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row, 1, f"Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")
    row += 2

    def _write_section(heading: str, data: dict):
        nonlocal row
        h = ws.cell(row, 1, heading.replace("_", " ").title())
        h.font = Font(bold=True, size=11)
        h.fill = PatternFill("solid", fgColor="1a1a25")
        h.font = Font(bold=True, color="FFFFFF")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        # header row
        for col, hdr in enumerate(["Field", "Value"], 1):
            c = ws.cell(row, col, hdr)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.border = thin
        row += 1

        for i, (k, v) in enumerate(data.items()):
            if isinstance(v, dict):
                continue
            fill = alt_fill if i % 2 == 0 else PatternFill()
            k_cell = ws.cell(row, 1, str(k).replace("_", " ").title())
            if isinstance(v, float):
                v_cell = ws.cell(row, 2, v)
                v_cell.number_format = '₹#,##0.00'
            else:
                v_cell = ws.cell(row, 2, str(v))
            k_cell.fill = fill
            v_cell.fill = fill
            k_cell.border = thin
            v_cell.border = thin
            row += 1
        row += 1

    for section_name, section_data in report_data.items():
        if isinstance(section_data, dict):
            _write_section(section_name, section_data)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _plain_text_report(data: dict, report_type: str) -> str:
    lines = [f"FinanceAI — {report_type.upper()} Report",
             f"Generated: {datetime.utcnow().isoformat()}", ""]
    for k, v in data.items():
        lines.append(f"{k}: {json.dumps(v, default=str, indent=2)}")
    return "\n".join(lines)
